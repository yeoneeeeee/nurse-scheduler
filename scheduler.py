import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from collections import defaultdict
import random

# =========================
# 파일 설정
# =========================
INPUT_FILE  = "template.xlsx"
OUTPUT_FILE = "schedule_output.xlsx"

# =========================
# 공휴일 설정
# =========================
HOLIDAYS = []

# =========================
# 비교대
# =========================
FIXED_D = ["문정희"]
FIXED_E = ["조복희"]

NON_SHIFT = FIXED_D + FIXED_E

# =========================
# N 금지자
# =========================
NO_NIGHT = []

# =========================
# 색상
# =========================
COLORS = {
    "D":   "ADD8E6",
    "SD":  "B0E0E6",
    "E":   "FFD580",
    "N":   "DDA0DD",
    "Off": "D9D9D9",
}

# =========================
# 필요 인원
# 비교대 제외 기준
# =========================
WEEKDAY_REQ = {"D": 8, "E": 8, "N": 5}
WEEKEND_REQ = {"D": 6, "E": 6, "N": 5}

# =========================
# 금지 패턴
# =========================
FORBIDDEN_NEXT = {
    ("E",  "D"),
    ("E",  "SD"),
    ("SD", "D"),
    ("N",  "D"),
    ("N",  "E"),
    ("N",  "SD"),
}

# =========================
# 엑셀 열기
# =========================
wb = load_workbook(INPUT_FILE)
ws = wb.active

NAME_START_ROW = 3
NAME_END_ROW   = 35

DATE_START_COL = 3
DATE_END_COL   = 33

TOTAL_DAYS = DATE_END_COL - DATE_START_COL + 1

# =========================
# 간호사 목록
# =========================
nurses = []
row_map = {}

for row in range(NAME_START_ROW, NAME_END_ROW + 1):
    name = ws.cell(row=row, column=2).value

    if name:
        nurses.append(name)
        row_map[name] = row

shift_nurses = [n for n in nurses if n not in NON_SHIFT]

# =========================
# 날짜 유틸
# =========================
def get_weekday_str(col):
    return ws.cell(row=2, column=col).value

def is_weekend_or_holiday(day_idx):
    col = DATE_START_COL + day_idx
    wd = get_weekday_str(col)

    day_num = day_idx + 1

    return wd in ("토", "일") or day_num in HOLIDAYS

def get_week(day_idx):

    col = DATE_START_COL + day_idx
    wd = get_weekday_str(col)

    WD_ORDER = {
        "일":0,
        "월":1,
        "화":2,
        "수":3,
        "목":4,
        "금":5,
        "토":6
    }

    sun_dist = WD_ORDER.get(wd, 0)

    week_start_idx = day_idx - sun_dist

    return week_start_idx // 7

# =========================
# 초기화
# =========================
schedule = defaultdict(dict)
night_count = defaultdict(int)

for name in nurses:
    for d in range(TOTAL_DAYS):
        schedule[name][d] = None

# =========================
# 비교대 먼저 고정
# =========================
for day_idx in range(TOTAL_DAYS):

    weekend = is_weekend_or_holiday(day_idx)

    for name in FIXED_D:

        if weekend:
            schedule[name][day_idx] = "Off"
        else:
            schedule[name][day_idx] = "D"

    for name in FIXED_E:

        if weekend:
            schedule[name][day_idx] = "Off"
        else:
            schedule[name][day_idx] = "E"

# =========================
# helper
# =========================
def prev(name, day_idx, n=1):
    return schedule[name].get(day_idx - n)

def work_streak_at(name, day_idx):

    cnt = 0
    idx = day_idx - 1

    while idx >= 0 and schedule[name].get(idx) not in ("Off", None):
        cnt += 1
        idx -= 1

    return cnt

def weekly_work_count(name, week):

    cnt = 0

    for d in range(TOTAL_DAYS):

        if get_week(d) == week:
            if schedule[name].get(d) not in ("Off", None):
                cnt += 1

    return cnt

# =========================
# 배정 가능 여부
# =========================
def can_assign(name, day_idx, shift):

    if schedule[name].get(day_idx) is not None:
        return False

    if shift == "N" and name in NO_NIGHT:
        return False

    if shift == "N" and night_count[name] >= 7:
        return False

    p1 = prev(name, day_idx, 1)
    p2 = prev(name, day_idx, 2)

    if p1 == "N" and shift != "Off":
        return False

    if p2 == "N" and p1 == "Off":
        if shift in ("D", "SD"):
            return False

    if (p1, shift) in FORBIDDEN_NEXT:
        return False

    if shift == "N":
        if p1 == "N" and p2 == "N":
            return False

    if shift != "Off" and work_streak_at(name, day_idx) >= 5:
        return False

    if shift != "Off":

        week = get_week(day_idx)

        if weekly_work_count(name, week) >= 5:
            return False

    return True

# =========================
# 메인 배정
# =========================
for day_idx in range(TOTAL_DAYS):

    req = WEEKEND_REQ.copy() if is_weekend_or_holiday(day_idx) else WEEKDAY_REQ.copy()

    # =====================
    # N 배정
    # =====================
    candidates = [
        n for n in shift_nurses
        if can_assign(n, day_idx, "N")
    ]

    candidates.sort(
        key=lambda n: (
            night_count[n],
            random.random()
        )
    )

    assigned = 0

    for name in candidates:

        if assigned >= req["N"]:
            break

        schedule[name][day_idx] = "N"

        night_count[name] += 1

        if day_idx + 1 < TOTAL_DAYS:
            schedule[name][day_idx + 1] = "Off"

        assigned += 1

    # =====================
    # D 배정
    # =====================
    candidates = [
        n for n in shift_nurses
        if can_assign(n, day_idx, "D")
    ]

    candidates.sort(
        key=lambda n: (
            sum(
                1 for d in range(TOTAL_DAYS)
                if schedule[n].get(d) == "D"
            ),
            random.random()
        )
    )

    assigned = 0

    for name in candidates:

        if assigned >= req["D"]:
            break

        schedule[name][day_idx] = "D"

        assigned += 1

    # =====================
    # E 배정
    # =====================
    candidates = [
        n for n in shift_nurses
        if can_assign(n, day_idx, "E")
    ]

    candidates.sort(
        key=lambda n: (
            sum(
                1 for d in range(TOTAL_DAYS)
                if schedule[n].get(d) == "E"
            ),
            random.random()
        )
    )

    assigned = 0

    for name in candidates:

        if assigned >= req["E"]:
            break

        schedule[name][day_idx] = "E"

        assigned += 1

    # =====================
    # 나머지 Off
    # =====================
    for name in shift_nurses:

        if schedule[name].get(day_idx) is None:
            schedule[name][day_idx] = "Off"

# =========================
# 검증 후 부족분 보정
# =========================
for day_idx in range(TOTAL_DAYS):

    req = WEEKEND_REQ if is_weekend_or_holiday(day_idx) else WEEKDAY_REQ

    counts = {"D":0, "E":0, "N":0}

    for name in shift_nurses:

        s = schedule[name][day_idx]

        if s in counts:
            counts[s] += 1

    # =====================
    # D 부족 보정
    # =====================
    while counts["D"] < req["D"]:

        changed = False

        for name in shift_nurses:

            if schedule[name][day_idx] == "Off":

                if can_assign(name, day_idx, "D"):

                    schedule[name][day_idx] = "D"

                    counts["D"] += 1

                    changed = True
                    break

        if not changed:
            break

    # =====================
    # E 부족 보정
    # =====================
    while counts["E"] < req["E"]:

        changed = False

        for name in shift_nurses:

            if schedule[name][day_idx] == "Off":

                if can_assign(name, day_idx, "E"):

                    schedule[name][day_idx] = "E"

                    counts["E"] += 1

                    changed = True
                    break

        if not changed:
            break

# =========================
# 엑셀 저장
# =========================
for name in nurses:

    row = row_map[name]

    for day_idx in range(TOTAL_DAYS):

        col = DATE_START_COL + day_idx

        shift = schedule[name][day_idx]

        cell = ws.cell(row=row, column=col)

        cell.value = shift

        color = COLORS.get(shift, "FFFFFF")

        cell.fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid"
        )

        cell.font = Font(
            name="맑은 고딕",
            size=9,
            bold=(shift == "N")
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )



# =========================
# 검증
# =========================
violations = []

for day_idx in range(TOTAL_DAYS):

    req = WEEKEND_REQ if is_weekend_or_holiday(day_idx) else WEEKDAY_REQ

    cnt = {"D":0, "E":0, "N":0}

    for name in shift_nurses:

        s = schedule[name][day_idx]

        if s in cnt:
            cnt[s] += 1

    for sh in ["D", "E", "N"]:

        if cnt[sh] != req[sh]:

            violations.append(
                f"{day_idx+1}일 {sh}: {cnt[sh]} / 필요 {req[sh]}"
            )

# =========================
# 저장
# =========================
if violations:
    print("❌ 조건 위반 발견. 파일 저장 안 함.")
    for v in violations:
        print(v)
    sys.exit(1)

else:
    wb.save(OUTPUT_FILE)
    print(f"✅ 모든 조건 만족. 저장 완료: {OUTPUT_FILE}")
    sys.exit(0)

print()

if violations:

    print("⚠️ 인원 조건 위반")

    for v in violations:
        print(v)

else:
    print("✅ 모든 조건 만족")
