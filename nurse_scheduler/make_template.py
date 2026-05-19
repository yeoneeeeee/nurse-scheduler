from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment

wb = Workbook()
ws = wb.active

# =========================
# 기본 스타일
# =========================
thin = Side(style='thin')

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

center = Alignment(horizontal='center', vertical='center')

# =========================
# 날짜 입력
# =========================
dates = list(range(1, 32))

col = 3  # C열 시작

for d in dates:
    cell = ws.cell(row=1, column=col)
    cell.value = d
    cell.border = border
    cell.alignment = center

    col += 1

# =========================
# 요일 입력 예시
# =========================
weekdays = [
    "월","화","수","목","금","토","일"
]

col = 3

for i in range(31):
    cell = ws.cell(row=2, column=col)
    cell.value = weekdays[i % 7]
    cell.border = border
    cell.alignment = center

    col += 1

# =========================
# 이름 영역
# =========================
names = [
    "문정희","조복희","백현경","김민영","선우현",
    "엄현정","서정현","김예진","박혜영","윤정윤",
    "한혜빈","이수현","정지수","조윤아","김별",
    "하소정","윤지인","김보현","정주리","서윤경",
    "박채원","이다영","박소정","유연화","김주현",
    "조황희","박혜윤","윤동영","이지영","강지은",
    "강미선","최유리","박예원"
]

row = 3

for name in names:

    # 이름
    name_cell = ws.cell(row=row, column=2)
    name_cell.value = name
    name_cell.border = border
    name_cell.alignment = center

    # 스케줄 영역
    for col in range(3, 34):
        c = ws.cell(row=row, column=col)
        c.border = border
        c.alignment = center

    row += 1

# =========================
# 저장
# =========================
wb.save("template.xlsx")

print("template.xlsx 생성 완료")